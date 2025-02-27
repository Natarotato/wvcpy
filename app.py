import tinytuya
import json
import base64
import struct
import time
import os
from datetime import datetime, time as dtime
from datetime import datetime, timedelta  # Add timedelta here
from openpyxl import Workbook, load_workbook
from typing import Dict, List, Optional, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import MaxNLocator
import threading
import pandas as pd
import random


# Load or initialize configuration (unchanged)
CONFIG_FILE = "config.json"
DEFAULT_CONFIG = {
    "API_KEY": "9kvfdejycye8958kh4ve",
    "API_SECRET": "759d51d9861c4ab68f0232836203ad2d",
    "REGION": "eu",
    "INVERTERS": [
        {"device_id": "bf53440f72897ffb99vda7", "ip": "172.17.4.184", "local_key": "_i^q|6=G2ghV69^a", "sheet": "Inverter 1"},
        {"device_id": "bf8d502c9fc3759ec6vgum", "ip": "172.17.6.176", "local_key": "|{#4oCF!URiNK3Ek", "sheet": "Inverter 2"},
        {"device_id": "", "ip": "", "local_key": "", "sheet": "Inverter 3"},
        {"device_id": "", "ip": "", "local_key": "", "sheet": "Inverter 4"}
    ],
    "RECORDING_WINDOW": {"start": "06:00", "stop": "20:00"},
    "FETCH_INTERVAL": 10,
    "SAVE_DIR": "data"
}

def load_config():
    default = DEFAULT_CONFIG.copy()
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            config = json.load(f)
            # Merge inverters, keeping existing ones and adding new ones if not present
            existing_ids = {inv["device_id"] for inv in config["INVERTERS"]}
            for default_inv in default["INVERTERS"]:
                if default_inv["device_id"] not in existing_ids:
                    config["INVERTERS"].append(default_inv)
            start_time = datetime.strptime(config["RECORDING_WINDOW"]["start"], "%H:%M").time()
            stop_time = datetime.strptime(config["RECORDING_WINDOW"]["stop"], "%H:%M").time()
            config["RECORDING_WINDOW"] = (start_time, stop_time)
            return config
    start_time = datetime.strptime(default["RECORDING_WINDOW"]["start"], "%H:%M").time()
    stop_time = datetime.strptime(default["RECORDING_WINDOW"]["stop"], "%H:%M").time()
    default["RECORDING_WINDOW"] = (start_time, stop_time)
    return default

CONFIG = load_config()

# Initialize Tuya Cloud (unchanged)
tinytuya.set_debug(False)
CLOUD = tinytuya.Cloud(
    apiRegion=CONFIG["REGION"],
    apiKey=CONFIG["API_KEY"],
    apiSecret=CONFIG["API_SECRET"],
    apiDeviceID=CONFIG["INVERTERS"][0]["device_id"]
)

# Functions (unchanged until fetch_inverter_data)
def decode_tuya_value(encoded_value: str) -> Optional[Tuple[int, ...]]:
    try:
        decoded_bytes = base64.b64decode(encoded_value)
        return struct.unpack(f">{len(decoded_bytes)//2}H", decoded_bytes)
    except Exception as e:
        print(f"⚠️ Base64 decoding error: {e}")
        return None

def fetch_inverter_data(device_id: str) -> Optional[Dict]:
    # Import app from the main module to access the InverterGUI instance
    try:
        from __main__ import app
        if app.simulate_mode and not device_id:  # Simulate for unconfigured inverters
            return {
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "important_dps": {
                    "reverse_energy_total (kWh)": random.uniform(0, 10),  # Random fake data
                    "temp_current (°C)": random.uniform(20, 30),
                    "ac_power (W)": random.uniform(50, 200)
                },
                "extracted": {
                    "phase_a": {
                        "ac_voltage": random.uniform(220, 240),
                        "frequency": random.uniform(49.9, 50.1),
                        "ac_current (A)": random.uniform(0.2, 1.0)
                    },
                    "pv1_dc_data": {
                        "dc_voltage": random.uniform(250, 350),
                        "dc_current": random.uniform(0.1, 0.5),
                        "dc_power": random.uniform(50, 150)
                    }
                }
            }
    except (ImportError, AttributeError):
        pass  # If app isn’t accessible, proceed to real fetch

    try:
        CLOUD.apiDeviceID = device_id
        status = CLOUD.getstatus(device_id)
        if not status or "result" not in status:
            print(f"❌ Failed to get status for device {device_id}")
            return None

        result = status["result"]
        data = result if isinstance(result, list) else [result]

        important_dps = {
            item["code"]: item["value"]
            for item in data
            if item.get("code", "") in ["reverse_energy_total", "temp_current", "ac_power"]
        }
        
        if "reverse_energy_total" in important_dps:
            important_dps["reverse_energy_total"] = important_dps["reverse_energy_total"] / 100
        if "ac_power" in important_dps:
            important_dps["ac_power"] = important_dps["ac_power"] / 10

        phase_a = next((decode_tuya_value(item["value"]) for item in data 
                       if item.get("code") == "phase_a"), None)
        pv1_dc_data = next((decode_tuya_value(item["value"]) for item in data 
                           if item.get("code") == "pv1_dc_data"), None)

        phase_a_data = {"ac_voltage": None, "frequency": None}
        if phase_a and len(phase_a) >= 2:
            phase_a_data = {"ac_voltage": phase_a[0] / 10, "frequency": phase_a[-1] / 10}

        pv1_dc_data_extracted = {"dc_voltage": None, "dc_current": None, "dc_power": None}
        if pv1_dc_data and len(pv1_dc_data) >= 3:
            pv1_dc_data_extracted = {
                "dc_voltage": pv1_dc_data[0] / 10,
                "dc_current": pv1_dc_data[1] / 10,
                "dc_power": pv1_dc_data[2] / 10
            }

        ac_current = (important_dps.get("ac_power") / phase_a_data["ac_voltage"] 
                     if phase_a_data["ac_voltage"] and "ac_power" in important_dps else None)

        return {
            "timestamp": datetime.now().strftime("%H:%M:%S"),
            "important_dps": {
                "reverse_energy_total (kWh)": important_dps.get("reverse_energy_total", "N/A"),
                "temp_current (°C)": important_dps.get("temp_current", "N/A"),
                "ac_power (W)": important_dps.get("ac_power", "N/A")
            },
            "extracted": {
                "phase_a": {**phase_a_data, "ac_current (A)": ac_current},
                "pv1_dc_data": pv1_dc_data_extracted
            }
        }
    except Exception as e:
        print(f"❌ Unexpected error for device {device_id}: {e}")
        return None

def write_to_excel(data: Dict, sheet_name: str, base_folder: str = CONFIG["SAVE_DIR"]) -> None:
    now = datetime.now()
    folder_path = os.path.join(base_folder, now.strftime("%Y-%m"))
    os.makedirs(folder_path, exist_ok=True)
    filename = os.path.join(folder_path, now.strftime("%Y-%m-%d.xlsx"))

    headers = ["Timestamp", "Reverse Energy (kWh)", "Temp (°C)", "AC Power (W)", "AC Voltage (V)", 
               "Frequency (Hz)", "AC Current (A)", "DC Voltage (V)", "DC Current (A)", "DC Power (W)"]

    wb = load_workbook(filename) if os.path.exists(filename) else Workbook()
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    
    if ws.max_row == 0 or (ws.max_row == 1 and not ws[1][0].value):
        ws.append(headers)

    ac_current = data["extracted"]["phase_a"].get("ac_current (A)")
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data["important_dps"]["reverse_energy_total (kWh)"],
        data["important_dps"]["temp_current (°C)"], data["important_dps"]["ac_power (W)"],
        data["extracted"]["phase_a"]["ac_voltage"], data["extracted"]["phase_a"]["frequency"],
        round(ac_current, 3) if isinstance(ac_current, (int, float)) else "N/A",
        data["extracted"]["pv1_dc_data"]["dc_voltage"], data["extracted"]["pv1_dc_data"]["dc_current"],
        data["extracted"]["pv1_dc_data"]["dc_power"]
    ]
    ws.append(row)
    wb.save(filename)
    print(f"✅ Data saved to '{filename}' in sheet '{sheet_name}'")

def load_historical_data(sheet_name: str, base_folder: str = CONFIG["SAVE_DIR"]) -> pd.DataFrame:
    now = datetime.now()
    folder = os.path.join(base_folder, now.strftime("%Y-%m"))
    all_data = []
    for file in os.listdir(folder):
        if file.endswith(".xlsx"):
            filename = os.path.join(folder, file)
            wb = load_workbook(filename)
            if sheet_name in wb.sheetnames:
                df = pd.read_excel(filename, sheet_name=sheet_name)
                df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors='coerce')
                all_data.append(df)
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    # ... (existing zero-fill logic for new sheets) ...
    filename = os.path.join(base_folder, now.strftime("%Y-%m"), now.strftime("%Y-%m-%d.xlsx"))
    headers = ["Timestamp", "Reverse Energy (kWh)", "Temp (°C)", "AC Power (W)", "AC Voltage (V)", 
               "Frequency (Hz)", "AC Current (A)", "DC Voltage (V)", "DC Current (A)", "DC Power (W)"]
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            df = pd.read_excel(filename, sheet_name=sheet_name)
            df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors='coerce')
            return df
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            zero_row = [now.strftime("%Y-%m-%d %H:%M:%S"), 0, 0, 0, 0, 0, 0, 0, 0, 0]
            ws.append(zero_row)
            wb.save(filename)
            df = pd.DataFrame([zero_row], columns=headers)
            df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors='coerce')
            return df
    else:
        wb = Workbook()
        ws = wb.active if sheet_name == "Sheet" else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(headers)
        zero_row = [now.strftime("%Y-%m-%d %H:%M:%S"), 0, 0, 0, 0, 0, 0, 0, 0, 0]
        ws.append(zero_row)
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        df = pd.DataFrame([zero_row], columns=headers)
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors='coerce')
        return df

class InverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Inverter Monitoring Dashboard")
        self.running = False
        self.last_update = "Never"
        self.resize_timer = None
        self.simulate_mode = False  # Add simulation mode flag

        # Main frame
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        # Tabs for inverters
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.grid(row=0, column=0, sticky="nsew")

        self.tabs = {}
        self.values = {}
        self.graphs = {}
        self.status_lights = {}

        for i, inverter in enumerate(CONFIG["INVERTERS"]):
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=inverter["sheet"])
            tab_id = inverter["device_id"] if inverter["device_id"] else f"unconfigured_{i}"
            self.tabs[tab_id] = tab
            self.values[tab_id] = {}
            self.status_lights[tab_id] = None
            self.setup_tab(tab, inverter["device_id"], inverter["sheet"], tab_id)  # Line 236

        # Control frame
        self.control_frame = ttk.Frame(self.main_frame)
        self.control_frame.grid(row=1, column=0, pady=5, sticky="ew")
        self.start_button = tk.Button(self.control_frame, text="Start", command=self.start_monitoring, bg="gray", fg="white")
        self.start_button.grid(row=0, column=0, padx=5)
        self.stop_button = tk.Button(self.control_frame, text="Stop", command=self.stop_monitoring, bg="gray", fg="white")
        self.stop_button.grid(row=0, column=1, padx=5)
        ttk.Button(self.control_frame, text="Refresh", command=self.refresh_data).grid(row=0, column=2, padx=5)
        ttk.Button(self.control_frame, text="Settings", command=self.open_settings).grid(row=0, column=3, padx=5)
        self.last_update_label = ttk.Label(self.control_frame, text=f"Last Update: {self.last_update}")
        self.last_update_label.grid(row=0, column=4, padx=5)
        ttk.Button(self.control_frame, text="Simulate", command=self.toggle_simulate).grid(row=0, column=5, padx=5)  # New button
        ttk.Button(self.control_frame, text="Export Data", command=self.export_historical_data).grid(row=0, column=6, padx=5)  # New button
        # Log at bottom
        self.log = tk.Text(self.main_frame, height=5, width=80, font=("Arial", 10))
        self.log.grid(row=2, column=0, pady=10, sticky="ew")

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.rowconfigure(0, weight=1)

        # Bind resize event with debouncing
        self.root.bind("<Configure>", self.on_resize)
    def toggle_simulate(self):
        self.simulate_mode = not self.simulate_mode
        self.log.insert(tk.END, f"[{datetime.now()}] Simulation {'enabled' if self.simulate_mode else 'disabled'}\n")
        # Update tab states for all inverters
        for i, inverter in enumerate(CONFIG["INVERTERS"]):
            tab_id = inverter["device_id"] if inverter["device_id"] else f"unconfigured_{i}"
            tab = self.tabs[tab_id]
            if not inverter["device_id"]:  # Only affect unconfigured inverters
                state = "normal" if self.simulate_mode else "disabled"
                self.notebook.tab(tab, state=state)
        if self.simulate_mode:
            self.refresh_data()  # Refresh data immediately when enabling simulation

            
    def export_historical_data(self):
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = "data/exports"
        os.makedirs(export_dir, exist_ok=True)
        for tab_id, graphs in self.graphs.items():
            historical_data = graphs["historical_data"]
            if not historical_data.empty:
                filename = f"{export_dir}/{tab_id}_historical_{now}.csv"
                historical_data.to_csv(filename, index=False)
                self.log.insert(tk.END, f"[{datetime.now()}] Exported historical data to {filename}\n")
    
    def setup_tab(self, tab, device_id, sheet_name, tab_id):
        data_frame = ttk.LabelFrame(tab, text="Current Values", padding="5")
        data_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        labels = ["AC Power (W)", "AC Voltage (V)", "Frequency (Hz)", "DC Power (W)", 
                  "DC Voltage (V)", "DC Current (A)", "Temperature (°C)", "Reverse Energy (kWh)"]
        for i, label in enumerate(labels):
            ttk.Label(data_frame, text=label, font=("Arial", 10)).grid(row=i, column=0, sticky="w", padx=5)
            value_label = ttk.Label(data_frame, text="N/A", font=("Arial", 10))
            value_label.grid(row=i, column=1, sticky="w", padx=5)
            self.values[tab_id][label] = value_label
        
        status_frame = ttk.Frame(data_frame)
        status_frame.grid(row=0, column=2, rowspan=len(labels), padx=5)
        ttk.Label(status_frame, text="Status", font=("Arial", 10)).grid(row=0, column=0)
        status_canvas = tk.Canvas(status_frame, width=20, height=20)
        status_canvas.grid(row=1, column=0)
        status_canvas.create_oval(2, 2, 18, 18, fill="grey" if device_id else "lightgrey", tags="status")
        self.status_lights[tab_id] = status_canvas
        
        if not device_id:
            self.notebook.tab(tab, state="disabled")

        graphs_frame = ttk.LabelFrame(tab, text="Trends", padding="5")
        graphs_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        tab.columnconfigure(1, weight=1)
        tab.rowconfigure(0, weight=1)
        graphs_frame.columnconfigure((0, 1), weight=1)
        graphs_frame.rowconfigure((0, 1), weight=1)

        range_frame = ttk.Frame(graphs_frame)
        range_frame.grid(row=0, column=0, columnspan=2, pady=5)
        self.range_var = tk.StringVar(value="All")
        ttk.OptionMenu(range_frame, self.range_var, "All", "Last Hour", "Last Day", "All", 
                       command=lambda val: self.update_all_graphs(tab_id)).pack()

        initial_width, initial_height = 8, 5

        power_frame = ttk.Frame(graphs_frame)
        power_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        power_select = tk.StringVar(value="Both")
        ttk.OptionMenu(power_frame, power_select, "Both", "AC", "DC", "Both", 
                       command=lambda val: self.update_power_graph(tab_id, val)).pack(fill="x")
        power_fig, power_ax = plt.subplots(figsize=(initial_width, initial_height))
        power_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
        power_canvas = FigureCanvasTkAgg(power_fig, master=power_frame)
        power_widget = power_canvas.get_tk_widget()
        power_widget.pack(fill="both", expand=True)

        voltage_frame = ttk.Frame(graphs_frame)
        voltage_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        voltage_select = tk.StringVar(value="Both")
        ttk.OptionMenu(voltage_frame, voltage_select, "Both", "AC", "DC", "Both", 
                       command=lambda val: self.update_voltage_graph(tab_id, val)).pack(fill="x")
        voltage_fig, voltage_ax = plt.subplots(figsize=(initial_width, initial_height))
        voltage_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
        voltage_canvas = FigureCanvasTkAgg(voltage_fig, master=voltage_frame)
        voltage_widget = voltage_canvas.get_tk_widget()
        voltage_widget.pack(fill="both", expand=True)

        current_frame = ttk.Frame(graphs_frame)
        current_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        current_select = tk.StringVar(value="Both")
        ttk.OptionMenu(current_frame, current_select, "Both", "AC", "DC", "Both", 
                       command=lambda val: self.update_current_graph(tab_id, val)).pack(fill="x")
        current_fig, current_ax = plt.subplots(figsize=(initial_width, initial_height))
        current_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
        current_canvas = FigureCanvasTkAgg(current_fig, master=current_frame)
        current_widget = current_canvas.get_tk_widget()
        current_widget.pack(fill="both", expand=True)

        energy_frame = ttk.Frame(graphs_frame)
        energy_frame.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
        ttk.Label(energy_frame, text="Reverse Energy").pack(fill="x")
        energy_fig, energy_ax = plt.subplots(figsize=(initial_width, initial_height))
        energy_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
        energy_canvas = FigureCanvasTkAgg(energy_fig, master=energy_frame)
        energy_widget = energy_canvas.get_tk_widget()
        energy_widget.pack(fill="both", expand=True)

        historical_data = load_historical_data(sheet_name)
        self.graphs[tab_id] = {
            "power_fig": power_fig, "power_ax": power_ax, "power_canvas": power_canvas, "power_select": power_select,
            "voltage_fig": voltage_fig, "voltage_ax": voltage_ax, "voltage_canvas": voltage_canvas, "voltage_select": voltage_select,
            "current_fig": current_fig, "current_ax": current_ax, "current_canvas": current_canvas, "current_select": current_select,
            "energy_fig": energy_fig, "energy_ax": energy_ax, "energy_canvas": energy_canvas,
            "historical_data": historical_data
        }
        self.update_all_graphs(tab_id)

    def setup_tab(self, tab, device_id, sheet_name, tab_id):
            data_frame = ttk.LabelFrame(tab, text="Current Values", padding="5")
            data_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

            labels = ["AC Power (W)", "AC Voltage (V)", "Frequency (Hz)", "DC Power (W)", 
                    "DC Voltage (V)", "DC Current (A)", "Temperature (°C)", "Reverse Energy (kWh)"]
            for i, label in enumerate(labels):
                ttk.Label(data_frame, text=label, font=("Arial", 10)).grid(row=i, column=0, sticky="w", padx=5)
                value_label = ttk.Label(data_frame, text="N/A", font=("Arial", 10))
                value_label.grid(row=i, column=1, sticky="w", padx=5)
                self.values[tab_id][label] = value_label  # Use tab_id instead of device_id
            
            status_frame = ttk.Frame(data_frame)
            status_frame.grid(row=0, column=2, rowspan=len(labels), padx=5)
            ttk.Label(status_frame, text="Status", font=("Arial", 10)).grid(row=0, column=0)
            status_canvas = tk.Canvas(status_frame, width=20, height=20)
            status_canvas.grid(row=1, column=0)
            status_canvas.create_oval(2, 2, 18, 18, fill="grey" if device_id else "lightgrey", tags="status")
            self.status_lights[tab_id] = status_canvas  # Use tab_id instead of device_id
            
            # Optionally grey out the tab for unconfigured inverters
            if not device_id:
                self.notebook.tab(tab, state="disabled")  # Disable tab until configured

            # Graphs frame with improved layout
            graphs_frame = ttk.LabelFrame(tab, text="Trends", padding="5")
            graphs_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
            tab.columnconfigure(1, weight=1)
            tab.rowconfigure(0, weight=1)
            graphs_frame.columnconfigure((0, 1), weight=1)
            graphs_frame.rowconfigure((0, 1), weight=1)

            # Range selector
            range_frame = ttk.Frame(graphs_frame)
            range_frame.grid(row=0, column=0, columnspan=2, pady=5)
            self.range_var = tk.StringVar(value="All")
            ttk.OptionMenu(range_frame, self.range_var, "All", "Last Hour", "Last Day", "All", 
                        command=lambda val: self.update_all_graphs(tab_id)).pack()

            # Initial figure size (will be adjusted on resize)
            initial_width, initial_height = 8, 5

            # Power Graph
            power_frame = ttk.Frame(graphs_frame)
            power_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
            power_select = tk.StringVar(value="Both")
            ttk.OptionMenu(power_frame, power_select, "Both", "AC", "DC", "Both", 
                        command=lambda val: self.update_power_graph(tab_id, val)).pack(fill="x")  # Use tab_id
            power_fig, power_ax = plt.subplots(figsize=(initial_width, initial_height))
            power_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)  # Set initial margins
            power_canvas = FigureCanvasTkAgg(power_fig, master=power_frame)
            power_widget = power_canvas.get_tk_widget()
            power_widget.pack(fill="both", expand=True)

            # Voltage Graph
            voltage_frame = ttk.Frame(graphs_frame)
            voltage_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
            voltage_select = tk.StringVar(value="Both")
            ttk.OptionMenu(voltage_frame, voltage_select, "Both", "AC", "DC", "Both", 
                        command=lambda val: self.update_voltage_graph(tab_id, val)).pack(fill="x")  # Use tab_id
            voltage_fig, voltage_ax = plt.subplots(figsize=(initial_width, initial_height))
            voltage_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
            voltage_canvas = FigureCanvasTkAgg(voltage_fig, master=voltage_frame)
            voltage_widget = voltage_canvas.get_tk_widget()
            voltage_widget.pack(fill="both", expand=True)

            # Current Graph
            current_frame = ttk.Frame(graphs_frame)
            current_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
            current_select = tk.StringVar(value="Both")
            ttk.OptionMenu(current_frame, current_select, "Both", "AC", "DC", "Both", 
                        command=lambda val: self.update_current_graph(tab_id, val)).pack(fill="x")  # Use tab_id
            current_fig, current_ax = plt.subplots(figsize=(initial_width, initial_height))
            current_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
            current_canvas = FigureCanvasTkAgg(current_fig, master=current_frame)
            current_widget = current_canvas.get_tk_widget()
            current_widget.pack(fill="both", expand=True)

            # Energy Graph
            energy_frame = ttk.Frame(graphs_frame)
            energy_frame.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
            ttk.Label(energy_frame, text="Reverse Energy").pack(fill="x")
            energy_fig, energy_ax = plt.subplots(figsize=(initial_width, initial_height))
            energy_fig.subplots_adjust(left=0.1, right=0.95, bottom=0.3, top=0.9)
            energy_canvas = FigureCanvasTkAgg(energy_fig, master=energy_frame)
            energy_widget = energy_canvas.get_tk_widget()
            energy_widget.pack(fill="both", expand=True)

            # Load initial historical data
            historical_data = load_historical_data(sheet_name)

            self.graphs[tab_id] = {
                "power_fig": power_fig, "power_ax": power_ax, "power_canvas": power_canvas, "power_select": power_select,
                "voltage_fig": voltage_fig, "voltage_ax": voltage_ax, "voltage_canvas": voltage_canvas, "voltage_select": voltage_select,
                "current_fig": current_fig, "current_ax": current_ax, "current_canvas": current_canvas, "current_select": current_select,
                "energy_fig": energy_fig, "energy_ax": energy_ax, "energy_canvas": energy_canvas,
                "historical_data": historical_data
            }

            # Initial plot
            self.update_all_graphs(tab_id)

    # Move update_power_graph outside setup_tab
    def update_power_graph(self, tab_id, option):
        graph_data = self.graphs[tab_id]
        historical_data = graph_data["historical_data"]
        graph_data["power_ax"].clear()
        if not historical_data.empty and pd.api.types.is_datetime64_any_dtype(historical_data["Timestamp"]):
            if self.range_var.get() == "Last Hour":
                historical_data = historical_data[historical_data["Timestamp"] > datetime.now() - timedelta(hours=1)]
            elif self.range_var.get() == "Last Day":
                historical_data = historical_data[historical_data["Timestamp"] > datetime.now() - timedelta(days=1)]
            time_only = historical_data["Timestamp"].dt.strftime("%H:%M:%S").fillna("N/A")
            if option in ["AC", "Both"] and not historical_data["AC Power (W)"].isna().all():
                graph_data["power_ax"].plot(time_only, historical_data["AC Power (W)"], 'b-', label="AC Power (W)")
            if option in ["DC", "Both"] and not historical_data["DC Power (W)"].isna().all():
                graph_data["power_ax"].plot(time_only, historical_data["DC Power (W)"], 'g-', label="DC Power (W)")
        graph_data["power_ax"].set_title("Power Trends")
        graph_data["power_ax"].legend(loc="upper left")
        graph_data["power_ax"].grid(True)
        graph_data["power_ax"].tick_params(axis='x', rotation=45)
        graph_data["power_ax"].xaxis.set_major_locator(MaxNLocator(nbins=6))
        graph_data["power_ax"].set_xlabel("Time (HH:MM:SS)")
        graph_data["power_fig"].tight_layout()
        graph_data["power_canvas"].draw()

    # ... (define update_voltage_graph, update_current_graph, update_energy_graph similarly) ...


    def on_resize(self, event):
        # Debounce the resize event
        if self.resize_timer is not None:
            self.root.after_cancel(self.resize_timer)
        self.resize_timer = self.root.after(200, self.resize_graphs)  # 200ms delay

    def resize_graphs(self):
        # Calculate new figure sizes based on graphs_frame size
        for device_id, tab in self.tabs.items():
            graphs_frame = tab.winfo_children()[1]  # Second child is graphs_frame
            width = graphs_frame.winfo_width() / 100   # Reduced divisor for larger width
            height = graphs_frame.winfo_height() / 200  # Reduced divisor for larger height
            graphs = self.graphs[device_id]
            for fig_key in ["power_fig", "voltage_fig", "current_fig", "energy_fig"]:
                fig = graphs[fig_key]
                new_width = max(4, width / 2 - 0.5)  # Increased min width
                new_height = max(3, height - 0.5)    # Increased min height
                fig.set_size_inches(new_width, new_height)
                # Full margin control: left, right, bottom, top (values between 0 and 1)
                fig.subplots_adjust(left=0.1, right=0.95, bottom=0.28, top=0.9)
                graphs[f"{fig_key.replace('_fig', '')}_canvas"].draw()

    def update_data(self):
        while self.running:
            current_time = datetime.now().time()
            if CONFIG["RECORDING_WINDOW"][0] <= current_time <= CONFIG["RECORDING_WINDOW"][1]:
                for i, inverter in enumerate(CONFIG["INVERTERS"]):
                    tab_id = inverter["device_id"] if inverter["device_id"] else f"unconfigured_{i}"
                    data = fetch_inverter_data(inverter["device_id"])
                    if data:
                        self.update_display(tab_id, data, inverter["sheet"])
                        write_to_excel(data, inverter["sheet"])
                        self.log.insert(tk.END, f"[{datetime.now()}] Data updated for {inverter['sheet']}\n")
                        self.status_lights[tab_id].itemconfig("status", fill="green")
                    else:
                        self.log.insert(tk.END, f"[{datetime.now()}] Failed to fetch data for {inverter['sheet']}\n")
                        self.status_lights[tab_id].itemconfig("status", fill="orange")
                    self.log.see(tk.END)
                    self.last_update = datetime.now().strftime("%H:%M:%S")
                    self.last_update_label.config(text=f"Last Update: {self.last_update}")
            time.sleep(CONFIG["FETCH_INTERVAL"])

    def refresh_data(self):
        for i, inverter in enumerate(CONFIG["INVERTERS"]):
            tab_id = inverter["device_id"] if inverter["device_id"] else f"unconfigured_{i}"
            data = fetch_inverter_data(inverter["device_id"])
            if data:
                self.update_display(tab_id, data, inverter["sheet"])
                self.status_lights[tab_id].itemconfig("status", fill="green")
            else:
                self.status_lights[tab_id].itemconfig("status", fill="orange")
            self.last_update = datetime.now().strftime("%H:%M:%S")
            self.last_update_label.config(text=f"Last Update: {self.last_update}")

    def update_display(self, device_id, data, sheet_name):
        def format_value(val):
            return f"{val:.2f}" if isinstance(val, (int, float)) else "N/A"

        self.values[device_id]["AC Power (W)"].config(text=format_value(data["important_dps"]["ac_power (W)"]))
        self.values[device_id]["AC Voltage (V)"].config(text=format_value(data["extracted"]["phase_a"]["ac_voltage"]))
        self.values[device_id]["Frequency (Hz)"].config(text=format_value(data["extracted"]["phase_a"]["frequency"]))
        self.values[device_id]["DC Power (W)"].config(text=format_value(data["extracted"]["pv1_dc_data"]["dc_power"]))
        self.values[device_id]["DC Voltage (V)"].config(text=format_value(data["extracted"]["pv1_dc_data"]["dc_voltage"]))
        self.values[device_id]["DC Current (A)"].config(text=format_value(data["extracted"]["pv1_dc_data"]["dc_current"]))
        temp = data["important_dps"]["temp_current (°C)"]
        self.values[device_id]["Temperature (°C)"].config(
            text=format_value(temp),
            foreground="red" if isinstance(temp, (int, float)) and temp > 50 else "black"
        )
        self.values[device_id]["Reverse Energy (kWh)"].config(text=format_value(data["important_dps"]["reverse_energy_total (kWh)"]))

        historical_data = self.graphs[device_id]["historical_data"]
        new_row = pd.DataFrame([{
            "Timestamp": datetime.now(),
            "Reverse Energy (kWh)": data["important_dps"]["reverse_energy_total (kWh)"],
            "Temp (°C)": data["important_dps"]["temp_current (°C)"],
            "AC Power (W)": data["important_dps"]["ac_power (W)"],
            "AC Voltage (V)": data["extracted"]["phase_a"]["ac_voltage"],
            "Frequency (Hz)": data["extracted"]["phase_a"]["frequency"],
            "AC Current (A)": data["extracted"]["phase_a"]["ac_current (A)"],
            "DC Voltage (V)": data["extracted"]["pv1_dc_data"]["dc_voltage"],
            "DC Current (A)": data["extracted"]["pv1_dc_data"]["dc_current"],
            "DC Power (W)": data["extracted"]["pv1_dc_data"]["dc_power"]
        }])
        self.graphs[device_id]["historical_data"] = pd.concat([historical_data, new_row], ignore_index=True)
        self.update_all_graphs(device_id)

    def update_all_graphs(self, device_id):
        self.update_power_graph(device_id, self.graphs[device_id]["power_select"].get())
        self.update_voltage_graph(device_id, self.graphs[device_id]["voltage_select"].get())
        self.update_current_graph(device_id, self.graphs[device_id]["current_select"].get())
        self.update_energy_graph(device_id)

    def update_power_graph(self, device_id, option):
        graph_data = self.graphs[device_id]
        historical_data = graph_data["historical_data"]
        graph_data["power_ax"].clear()
        # Check if Timestamp is datetime-like and not empty
        if not historical_data.empty and pd.api.types.is_datetime64_any_dtype(historical_data["Timestamp"]):
            time_only = historical_data["Timestamp"].dt.strftime("%H:%M:%S").fillna("N/A")
        else:
            time_only = []  # Empty list for no data
        if option in ["AC", "Both"] and not historical_data["AC Power (W)"].isna().all():
            graph_data["power_ax"].plot(time_only, historical_data["AC Power (W)"], 'b-', label="AC Power (W)")
        if option in ["DC", "Both"] and not historical_data["DC Power (W)"].isna().all():
            graph_data["power_ax"].plot(time_only, historical_data["DC Power (W)"], 'g-', label="DC Power (W)")
        graph_data["power_ax"].set_title("Power Trends")
        graph_data["power_ax"].legend(loc="upper left")
        graph_data["power_ax"].grid(True)
        graph_data["power_ax"].tick_params(axis='x', rotation=45)
        graph_data["power_ax"].xaxis.set_major_locator(MaxNLocator(nbins=6))
        graph_data["power_ax"].set_xlabel("Time (HH:MM:SS)")
        graph_data["power_fig"].tight_layout()
        graph_data["power_canvas"].draw()

    def update_voltage_graph(self, device_id, option):
        graph_data = self.graphs[device_id]
        historical_data = graph_data["historical_data"]
        graph_data["voltage_ax"].clear()
        if not historical_data.empty and pd.api.types.is_datetime64_any_dtype(historical_data["Timestamp"]):
            time_only = historical_data["Timestamp"].dt.strftime("%H:%M:%S").fillna("N/A")
        else:
            time_only = []
        if option in ["AC", "Both"] and not historical_data["AC Voltage (V)"].isna().all():
            graph_data["voltage_ax"].plot(time_only, historical_data["AC Voltage (V)"], 'b-', label="AC Voltage (V)")
        if option in ["DC", "Both"] and not historical_data["DC Voltage (V)"].isna().all():
            graph_data["voltage_ax"].plot(time_only, historical_data["DC Voltage (V)"], 'g-', label="DC Voltage (V)")
        graph_data["voltage_ax"].set_title("Voltage Trends")
        graph_data["voltage_ax"].legend(loc="upper left")
        graph_data["voltage_ax"].grid(True)
        graph_data["voltage_ax"].tick_params(axis='x', rotation=45)
        graph_data["voltage_ax"].xaxis.set_major_locator(MaxNLocator(nbins=6))
        graph_data["voltage_ax"].set_xlabel("Time (HH:MM:SS)")
        graph_data["voltage_fig"].tight_layout()
        graph_data["voltage_canvas"].draw()

    def update_current_graph(self, device_id, option):
        graph_data = self.graphs[device_id]
        historical_data = graph_data["historical_data"]
        graph_data["current_ax"].clear()
        if not historical_data.empty and pd.api.types.is_datetime64_any_dtype(historical_data["Timestamp"]):
            time_only = historical_data["Timestamp"].dt.strftime("%H:%M:%S").fillna("N/A")
        else:
            time_only = []
        if option in ["AC", "Both"] and not historical_data["AC Current (A)"].isna().all():
            graph_data["current_ax"].plot(time_only, historical_data["AC Current (A)"], 'b-', label="AC Current (A)")
        if option in ["DC", "Both"] and not historical_data["DC Current (A)"].isna().all():
            graph_data["current_ax"].plot(time_only, historical_data["DC Current (A)"], 'g-', label="DC Current (A)")
        graph_data["current_ax"].set_title("Current Trends")
        graph_data["current_ax"].legend(loc="upper left")
        graph_data["current_ax"].grid(True)
        graph_data["current_ax"].tick_params(axis='x', rotation=45)
        graph_data["current_ax"].xaxis.set_major_locator(MaxNLocator(nbins=6))
        graph_data["current_ax"].set_xlabel("Time (HH:MM:SS)")
        graph_data["current_fig"].tight_layout()
        graph_data["current_canvas"].draw()

    def update_energy_graph(self, device_id):
        graph_data = self.graphs[device_id]
        historical_data = graph_data["historical_data"]
        graph_data["energy_ax"].clear()
        if not historical_data.empty and pd.api.types.is_datetime64_any_dtype(historical_data["Timestamp"]):
            time_only = historical_data["Timestamp"].dt.strftime("%H:%M:%S").fillna("N/A")
        else:
            time_only = []
        if not historical_data["Reverse Energy (kWh)"].isna().all():
            graph_data["energy_ax"].plot(time_only, historical_data["Reverse Energy (kWh)"], 'm-', label="Energy (kWh)")
        graph_data["energy_ax"].set_title("Energy Trends")
        graph_data["energy_ax"].legend(loc="upper left")
        graph_data["energy_ax"].grid(True)
        graph_data["energy_ax"].tick_params(axis='x', rotation=45)
        graph_data["energy_ax"].xaxis.set_major_locator(MaxNLocator(nbins=6))
        graph_data["energy_ax"].set_xlabel("Time (HH:MM:SS)")
        graph_data["energy_fig"].tight_layout()
        graph_data["energy_canvas"].draw()

    def start_monitoring(self):
        if not self.running:
            self.running = True
            self.thread = threading.Thread(target=self.update_data, daemon=True)
            self.thread.start()
            self.log.insert(tk.END, f"[{datetime.now()}] Monitoring started\n")
            self.start_button.config(bg="green")
            self.stop_button.config(bg="gray")

    def stop_monitoring(self):
        if self.running:
            self.running = False
            self.log.insert(tk.END, f"[{datetime.now()}] Monitoring stopped\n")
            self.start_button.config(bg="gray")
            self.stop_button.config(bg="red")

    def open_settings(self):
        settings_win = tk.Toplevel(self.root)
        settings_win.title("Settings")
        settings_win.geometry("400x700")

        api_frame = ttk.LabelFrame(settings_win, text="API Settings", padding="5")
        api_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(api_frame, text="API Key:").grid(row=0, column=0, sticky="w")
        api_key_entry = ttk.Entry(api_frame, width=30)
        api_key_entry.insert(0, CONFIG["API_KEY"])
        api_key_entry.grid(row=0, column=1)
        ttk.Label(api_frame, text="API Secret:").grid(row=1, column=0, sticky="w")
        api_secret_entry = ttk.Entry(api_frame, width=30)
        api_secret_entry.insert(0, CONFIG["API_SECRET"])
        api_secret_entry.grid(row=1, column=1)
        ttk.Label(api_frame, text="Region:").grid(row=2, column=0, sticky="w")
        region_entry = ttk.Entry(api_frame, width=30)
        region_entry.insert(0, CONFIG["REGION"])
        region_entry.grid(row=2, column=1)

        recording_frame = ttk.LabelFrame(settings_win, text="Recording Settings", padding="5")
        recording_frame.pack(fill="x", padx=5, pady=5)
        ttk.Label(recording_frame, text="Start Time (HH:MM):").grid(row=0, column=0, sticky="w")
        start_time_entry = ttk.Entry(recording_frame, width=10)
        start_time_entry.insert(0, CONFIG["RECORDING_WINDOW"][0].strftime("%H:%M"))
        start_time_entry.grid(row=0, column=1, sticky="w")
        ttk.Label(recording_frame, text="Stop Time (HH:MM):").grid(row=1, column=0, sticky="w")
        stop_time_entry = ttk.Entry(recording_frame, width=10)
        stop_time_entry.insert(0, CONFIG["RECORDING_WINDOW"][1].strftime("%H:%M"))
        stop_time_entry.grid(row=1, column=1, sticky="w")
        ttk.Label(recording_frame, text="Interval (seconds):").grid(row=2, column=0, sticky="w")
        interval_entry = ttk.Entry(recording_frame, width=10)
        interval_entry.insert(0, str(CONFIG["FETCH_INTERVAL"]))
        interval_entry.grid(row=2, column=1, sticky="w")

        inverter_frame = ttk.LabelFrame(settings_win, text="Inverters", padding="5")
        inverter_frame.pack(fill="x", padx=5, pady=5)
        inverter_entries = []
        for i, inv in enumerate(CONFIG["INVERTERS"]):
            ttk.Label(inverter_frame, text=f"Inverter {i+1}").grid(row=i*5, column=0, columnspan=2)
            entries = {}
            for j, (key, label) in enumerate([("device_id", "Device ID:"), ("ip", "IP:"), ("local_key", "Local Key:"), ("sheet", "Sheet Name:")]):
                ttk.Label(inverter_frame, text=label).grid(row=i*5+j+1, column=0, sticky="w")
                entry = ttk.Entry(inverter_frame, width=30)
                entry.insert(0, inv[key])
                entry.grid(row=i*5+j+1, column=1)
                entries[key] = entry
            inverter_entries.append(entries)

        save_frame = ttk.LabelFrame(settings_win, text="Data Save Location", padding="5")
        save_frame.pack(fill="x", padx=5, pady=5)
        save_dir_entry = ttk.Entry(save_frame, width=30)
        save_dir_entry.insert(0, CONFIG["SAVE_DIR"])
        save_dir_entry.grid(row=0, column=0)
        ttk.Button(save_frame, text="Browse", command=lambda: save_dir_entry.insert(0, filedialog.askdirectory())).grid(row=0, column=1)

        ttk.Button(settings_win, text="Save", command=lambda: self.save_settings(
            api_key_entry.get(), api_secret_entry.get(), region_entry.get(), 
            start_time_entry.get(), stop_time_entry.get(), interval_entry.get(),
            inverter_entries, save_dir_entry.get())).pack(pady=10)

    def save_settings(self, api_key, api_secret, region, start_time, stop_time, interval, inverter_entries, save_dir):
        global CONFIG, CLOUD
        try:
            start = datetime.strptime(start_time, "%H:%M").time()
            stop = datetime.strptime(stop_time, "%H:%M").time()
            interval_int = int(interval)
            if interval_int <= 0:
                raise ValueError("Interval must be positive")
        except ValueError as e:
            messagebox.showerror("Invalid Input", f"Error: {e}. Use HH:MM for times and a positive integer for interval.")
            return

        CONFIG["API_KEY"] = api_key
        CONFIG["API_SECRET"] = api_secret
        CONFIG["REGION"] = region
        CONFIG["RECORDING_WINDOW"] = (start, stop)
        CONFIG["FETCH_INTERVAL"] = interval_int
        CONFIG["INVERTERS"] = [
            {key: entries[key].get() for key in ["device_id", "ip", "local_key", "sheet"]}
            for entries in inverter_entries
        ]
        CONFIG["SAVE_DIR"] = save_dir

        config_to_save = CONFIG.copy()
        config_to_save["RECORDING_WINDOW"] = {
            "start": start.strftime("%H:%M"),
            "stop": stop.strftime("%H:%M")
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config_to_save, f, indent=4)

        CLOUD = tinytuya.Cloud(
            apiRegion=CONFIG["REGION"],
            apiKey=CONFIG["API_KEY"],
            apiSecret=CONFIG["API_SECRET"],
            apiDeviceID=CONFIG["INVERTERS"][0]["device_id"]
        )
        self.log.insert(tk.END, f"[{datetime.now()}] Configuration updated\n")
        if self.running:
            self.refresh_data()

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1400x900")
    app = InverterGUI(root)
    root.mainloop()